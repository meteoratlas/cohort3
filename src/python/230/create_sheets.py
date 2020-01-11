import random
from datetime import date
from openpyxl import Workbook

# from http://random-name-generator.info/, which generates exclusively old white person names
names = ["Mindy Thorson", "Chong Zartman", "Helena Treiber",
"Lorenzo Beaton", "Dorethea Clapper", "Stanford Woll",
"Sandy Greenman", "Aurora Rohlfing", "Austin Audette",
"Delbert Hoey", "Janie Valenzula"]

products = [("Tenor Saxophone", 1299.34), ("Electric Guitar", 1075.50), ("Banjo", 356.00), ("Clarinet", 595.99), ("MS-20", 650.30), ("Cello", 1495.50), ("Tuba", 3500.99)]
purchases = []

def rand_phone_num():
    r = random.randint
    return f"{r(300,999)}-{r(100,999)}-{r(1000,9999)}"

def random_date():
    start = date.today().replace(day=1, month=1, year=2019).toordinal()
    end = date.today().toordinal()
    return date.fromordinal(random.randint(start, end))

def make_email(first, last):
    domains = ["yahoo.org", "aol.ca", "gmail.com", "face.book", "microsoft.ls"]
    return f"{first.lower()}.{last.lower()}@{random.choice(domains)}"

def create_customer_table(ws, names):
    fields = ["customer_id", "first_name", "last_name", "phone_number", "email"]
    ws.append(fields)
    for i,name in enumerate(names):
        first, last = name.split()
        ws.append([i, first, last, rand_phone_num(), make_email(first, last)])

def create_product_table(ws):
    fields = ["product_id", "name", "price"]
    ws.append(fields)
    for i,product in enumerate(products):
        instrument, cost = product
        ws.append([i, instrument, cost])

def create_invoice_item_table(ws, amount):
    fields = ["invoice_item_id", "product_id", "quantity"]
    ws.append(fields)
    for i in range(amount):
        purchases.append((random.randint(0, len(products)), random.randint(1,4)))
        ws.append([i, purchases[i][0], purchases[i][1]])

def create_invoice_table(ws, amount):
    fields = ["invoice_id", "invoice_item_id", "customer_id", "date", "total_cost", "paid"]
    ws.append(fields)

    for i in range(amount):
        product_id, quantity = purchases[i]
        ws.append([i, i, random.randint(0,len(names)), random_date(), products[product_id][1] * quantity, random.choice([True, False]) ])

def generate():
    wb = Workbook()
    filename = "data.xlsx"
    
    customer_ws = wb.active
    customer_ws.title = "Customers"
    create_customer_table(customer_ws, names)

    products_ws = wb.create_sheet(title="Products")
    create_product_table(products_ws)

    num_of_invoice_items = 12
    invoice_items_ws = wb.create_sheet(title="Invoice Items")
    create_invoice_item_table(invoice_items_ws, num_of_invoice_items)

    num_of_invoices = 10
    invoices_ws = wb.create_sheet(title="Invoices")
    create_invoice_table(invoices_ws, num_of_invoices)

    wb.save(filename = filename)

if __name__ == "__main__":
    generate()