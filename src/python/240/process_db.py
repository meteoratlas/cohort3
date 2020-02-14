from openpyxl import Workbook, load_workbook
import datetime
import json

def read_workbook(file, flattenBools = True):
    def flattenBool(b):
        if flattenBools:
            if b == "=TRUE()": return True
            elif b == "=FALSE()": return False
        return b
        
    book = load_workbook(file)
    sheet_dicts = {}
    for sheet in book:
        data = []
        titles = [t.value for t in sheet[1]]
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            # Skip column header row 
            if (i == 0):
                continue
            row_data = {}
            for j, info in enumerate(row):
                row_data[titles[j]] = flattenBool(info)
            data.append(row_data)
        sheet_dicts[sheet.title] = data
    return sheet_dicts

def make_JSON(data):
    def dateToString(date):
        if isinstance(date, datetime.datetime):
            return date.__str__()

    with open("info.json", "w") as txt:
        txt.write(json.dumps(data, default = dateToString, indent=4, separators=(',', ': ')))